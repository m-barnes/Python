import openpyxl
import sys

from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side

#Make sure only one integer argument is passed from the command prompt. If it isn't, provide an error message and exit the script.
if len(sys.argv)!=2:
	print('multiplication_table.py requires a single integer value. Please try again.')
	sys.exit()
#placeholder variable for the number passed.
number = None
#Two system arguments are passed. The name of the file and the number. Grab the second value in the argument list, recast it as an integer, and make sure it's positive. As you can't create multiplcation tables of decimal or negative numbers. If it can't, provide an error message and exit the script. 
try:
	number = abs(int(sys.argv[1]))
except ValueError as e:
	print('That is not a number. Please try again.')
	sys.exit(1)

#set a variable, use openpyxl to create a workbook. Set the current sheet.	
wb = openpyxl.Workbook()
sheet = wb.active

#add a bunch of styling properties. 
bold_white = Font(name='Times New Roman', bold = True, color = 'ffffff')
title = Font(size = 24, italic = True)
blue_fill = PatternFill(start_color = '006294', end_color='006294', fill_type='solid')
centered_text = Alignment(horizontal = 'center', vertical = 'center', text_rotation=0, shrink_to_fit = True, indent = 0)
border = Border(left=Side(style='thin'), right=Side(style='thin'),top = Side(style='thin'), bottom=Side(style='thin'))

#Grab the number and add two to create a table of the right size. Start filling in values.
for row_num in range(1, number+2):
	for col_num in range(1, number+2):
		#if the cell is A1, do nothing. Color it blue and fill it empty space.		
		if row_num==1 and col_num==1:
			sheet.cell(row=row_num, column=col_num).value=''
			sheet.cell(row=row_num, column=col_num).fill = blue_fill
		#if row number equals 1, style the cell. 	
		elif row_num==1:
			sheet.cell(row=row_num, column=col_num).value = col_num-1
			sheet.cell(row=row_num, column=col_num).font = bold_white
			sheet.cell(row=row_num, column=col_num).fill = blue_fill
			sheet.cell(row=row_num, column=col_num).alignment = centered_text
			sheet.row_dimensions[row_num].height = 35
		#if the column number equals 1, style the cell. 
		elif col_num==1:
			sheet.cell(row=row_num, column=col_num).value=row_num-1
			sheet.cell(row=row_num, column=col_num).font =  bold_white
			sheet.cell(row=row_num, column=col_num).fill = blue_fill
			sheet.cell(row=row_num, column=col_num).alignment = centered_text
			sheet.row_dimensions[row_num].height = 35
		#if the row != 1 and column != 1 fill it in with a value accordingly, style the cell. 	
		else:
			sheet.cell(row=row_num, column=col_num).value = (row_num-1)*(col_num-1)
			sheet.cell(row=row_num, column=col_num).border = border
			sheet.cell(row=row_num, column=col_num).alignment = centered_text
			
#At the end, take the last row of cells and merge them together. Add a title and style the new merged cell. 			
sheet.merge_cells(start_row= int(row_num+1), start_column = 1, end_row = int(row_num+1), end_column=int(col_num))
sheet.cell(row=row_num + 1, column=1).value = ('Multiplication Table {0}\'s'.format(number))
sheet.cell(row=row_num + 1, column=1).font = title
sheet.cell(row=row_num + 1, column=1).alignment = centered_text
sheet.row_dimensions[row_num+1].height=70

#Save the table as a spreadsheet. Use the number argument passed to title the file. Exit the script. 
wb.save('Multiplication_Table_{0}\'s.xlsx'.format(number))
print('File created successfully!')
sys.exit()
