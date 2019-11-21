import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers={}

for r in range(2, sheet.max_row +1):
	payment = sheet.cell(row=r, column = lastCol).value
	if payment!='paid':
		name = sheet.cell(row=r, column = 1).value
		email = sheet.cell(row=r, column = 2).value
		unpaidMembers[name] = email
		
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('m.brandon.barnes@gmail.com', 'xxxxxxxxxxxxxx')

for name, email in unpaidMembers.items():
	body = 'Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Pay up chump. Thanks!' % (latestMonth, name, latestMonth)
	
	print('Sending email to %s...' % email)
	sendmailStatus = smtpObj.sendmail('xxxxxxxxxxxx@gmail.com', email, body)
	
	if sendmailStatus != {}:
		print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()
	
